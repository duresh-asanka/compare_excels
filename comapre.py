from openpyxl import load_workbook
from colorama import Fore, Back, Style


local = load_workbook('local.xlsx')
remote = load_workbook('remote.xlsx')

sheetnames = local.sheetnames

#iterate through each sheet
for sheets in sheetnames:
	local_sheet = local[sheets]
	remote_sheet = remote[sheets]
	#now compare each values from each sheet
	local_sheet_values = []
	remote_sheet_values = []
	#fill the list with the values in the local file of the current selected sheet
	for row in local_sheet.values:
		for value in row:
			local_sheet_values.append(value)
	#fill the list with the values in the remote file of the current selected sheet
	for row in remote_sheet.values:
		for value in row:
			remote_sheet_values.append(value)
	#now compare two lists
	if local_sheet_values == remote_sheet_values:
		print (local_sheet , Back.GREEN + ' --EQUIVALENT-- ' , Style.RESET_ALL)
	else:
		print (local_sheet, Back.RED + ' --DIFFERENT-- ', Style.RESET_ALL)


